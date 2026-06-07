"""
Eksport wyczyszczonych danych i raportu jakosci do pliku Excel.

Generuje sformatowany skoroszyt z zakladkami: dane, podsumowanie sprzedazy,
raport jakosci, wartosci do recznej poprawy i auto-poprawki fuzzy. Potrafi
zapisac na dysk (sciezka) lub do pamieci (BytesIO) - to drugie dla pobierania
raportu z aplikacji webowej.
"""
import io
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

KOLUMNY_DANE = ["data_sprzedazy", "region", "sprzedawca", "produkt",
                "ilosc", "cena_jednostkowa", "wartosc", "plik_zrodlowy", "potencjalny_duplikat"]

NAGLOWEK_FILL = PatternFill("solid", start_color="1F4E78")
NAGLOWEK_FONT = Font(name="Arial", bold=True, color="FFFFFF")
ZWYKLY_FONT = Font(name="Arial")
TYTUL_FONT = Font(name="Arial", bold=True, size=12)


def _zbuduj_jakosc(raport):
    fuzzy_n = len(raport.get("region_fuzzy", [])) + len(raport.get("produkt_fuzzy", []))
    wiersze = [
        ("Daty znormalizowane (inny format)", raport["daty_przeksztalcone"], "Sprowadzone do jednego formatu"),
        ("Daty niesparsowane", raport["daty_niesparsowane"], "Nie dalo sie odczytac - do sprawdzenia"),
        ("Ceny naprawione (przecinek)", raport["cena_przecinek"], "'4,29' -> 4.29"),
        ("Ceny puste", raport["cena_pusta"], "Brak ceny - wartosc nie policzona"),
        ("Ilosci ujemne naprawione", raport["ilosc_ujemna"], "Zalozono literowke znaku (abs)"),
        ("Sprzedawcy uzupelnieni", raport["sprzedawca_brak"], "Puste pola -> 'Nieznany'"),
        ("Wartosci przeliczone", raport["wartosc_niezgodna"], "wartosc != ilosc*cena -> przeliczone"),
        ("Wartosci bez ceny", raport["wartosc_bez_ceny"], "Brak ceny -> wartosc pusta"),
        ("Duplikaty do weryfikacji", raport["duplikaty_do_weryfikacji"], "NIE usuniete - sprawdzic recznie"),
        ("Auto-poprawki fuzzy", fuzzy_n, "Literowki dopasowane automatycznie - sprawdz w zakladce Auto_poprawione"),
    ]
    return pd.DataFrame(wiersze, columns=["Kategoria", "Liczba", "Opis"])


def _zbuduj_auto_poprawione(raport):
    # Lista auto-poprawek fuzzy do AUDYTU przez czlowieka. To zgadywanie
    # (literowka dopasowana do znanej nazwy), wiec pokazujemy co i z jaka
    # pewnoscia zostalo zmienione, zeby dalo sie zweryfikowac lub cofnac.
    wiersze = []
    for kolumna, klucz in [("region", "region_fuzzy"), ("produkt", "produkt_fuzzy")]:
        for p in raport.get(klucz, []):
            wiersze.append({
                "Kolumna": kolumna,
                "Bylo": p["wartosc"],
                "Poprawiono na": p["poprawiono_na"],
                "Podobienstwo": p["wynik"],
            })
    if not wiersze:
        return pd.DataFrame([{"Kolumna": "-", "Bylo": "(brak)", "Poprawiono na": "-",
                              "Podobienstwo": "Nic nie auto-poprawiono"}])
    return pd.DataFrame(wiersze)


def _zbuduj_do_poprawienia(raport):
    wszystkie = raport["region_nierozpoznane"] + raport["produkt_nierozpoznane"]
    if not wszystkie:
        return pd.DataFrame([{"Wartosc": "(brak)", "Plik": "-", "Liczba": 0,
                              "Instrukcja": "Wszystkie wartosci rozpoznane - nic do poprawy"}])
    return pd.DataFrame([{
        "Wartosc": r["wartosc"], "Plik": r["plik"], "Liczba": r["liczba"],
        "Instrukcja": f"Otworz '{r['plik']}', znajdz '{r['wartosc']}' (Ctrl+F) i popraw.",
    } for r in wszystkie])


def _formatuj_naglowek(ws):
    for kom in ws[1]:
        kom.font = NAGLOWEK_FONT
        kom.fill = NAGLOWEK_FILL
        kom.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _dopasuj_szerokosci(ws):
    for kolumna in ws.columns:
        dl = max((len(str(k.value)) for k in kolumna if k.value is not None), default=10)
        ws.column_dimensions[kolumna[0].column_letter].width = min(dl + 3, 50)


def eksportuj_do_excela(df, raport, sciezka="raporty/raport_sprzedaz.xlsx"):
    """Zapisuje dane i raport jakosci do sformatowanego skoroszytu Excel.

    sciezka moze byc tekstem (zapis na dysk) lub buforem BytesIO (zapis do
    pamieci - do pobrania w aplikacji webowej). Zwraca uzyta sciezke/bufor.
    """
    # Folder tworzymy tylko gdy zapisujemy na DYSK (sciezka to tekst).
    # Gdy sciezka to bufor w pamieci (BytesIO) - pomijamy, bufor nie ma folderu.
    if isinstance(sciezka, str):
        os.makedirs(os.path.dirname(sciezka), exist_ok=True)

    # agregaty - liczone w pandas (to raport-zrzut, nie model do edycji w Excelu)
    agregaty = {
        "Sprzedaz wg miesiaca": df.assign(m=df["data_sprzedazy"].dt.to_period("M").astype(str))
                                  .groupby("m")["wartosc"].sum().round(2),
        "Sprzedaz wg regionu": df.groupby("region")["wartosc"].sum().round(2).sort_values(ascending=False),
        "Sprzedaz wg produktu": df.groupby("produkt")["wartosc"].sum().round(2).sort_values(ascending=False),
        "Sprzedaz wg sprzedawcy": df.groupby("sprzedawca")["wartosc"].sum().round(2).sort_values(ascending=False),
    }

    # 1. Dane + tabele raportowe (pandas)
    with pd.ExcelWriter(sciezka, engine="openpyxl") as writer:
        df[KOLUMNY_DANE].to_excel(writer, sheet_name="Dane", index=False)
        _zbuduj_jakosc(raport).to_excel(writer, sheet_name="Jakosc_danych", index=False)
        _zbuduj_do_poprawienia(raport).to_excel(writer, sheet_name="Do_poprawienia", index=False)
        _zbuduj_auto_poprawione(raport).to_excel(writer, sheet_name="Auto_poprawione", index=False)
    # 2. Formatowanie + zakladka Podsumowanie (openpyxl)
    wb = load_workbook(sciezka)

    # Podsumowanie budujemy recznie, zeby miec czyste tytuly bez dublowania naglowkow
    ps = wb.create_sheet("Podsumowanie", index=1)  # jako druga zakladka
    r = 1
    for tytul, seria in agregaty.items():
        ps.cell(row=r, column=1, value=tytul).font = TYTUL_FONT
        r += 1
        naglowki = ps.cell(row=r, column=1, value="Kategoria"), ps.cell(row=r, column=2, value="Wartosc (zl)")
        for kom in naglowki:
            kom.font = NAGLOWEK_FONT
            kom.fill = NAGLOWEK_FILL
        r += 1
        for etykieta, wartosc in seria.items():
            ps.cell(row=r, column=1, value=etykieta)
            kom_w = ps.cell(row=r, column=2, value=float(wartosc))
            kom_w.number_format = '#,##0.00 "zl"'
            r += 1
        r += 1  # pusty wiersz miedzy tabelami

    # czcionka Arial wszedzie + naglowki/szerokosci
    for ws in wb.worksheets:
        for wiersz in ws.iter_rows():
            for kom in wiersz:
                if not (kom.font and kom.font.bold):
                    kom.font = ZWYKLY_FONT
        _dopasuj_szerokosci(ws)
    for nazwa in ("Dane", "Jakosc_danych", "Do_poprawienia", "Auto_poprawione"):
        _formatuj_naglowek(wb[nazwa])

    # Dane: format daty i waluty + filtr
    dane = wb["Dane"]
    for wiersz in dane.iter_rows(min_row=2):
        wiersz[0].number_format = "YYYY-MM-DD"
        wiersz[5].number_format = '#,##0.00 "zl"'
        wiersz[6].number_format = '#,##0.00 "zl"'
    dane.auto_filter.ref = dane.dimensions

    # Gdy zapisujemy do bufora w pamieci (BytesIO), ten sam bufor zostal juz
    # raz zapisany przez ExcelWriter wyzej. Drugi zapis NIE kasuje pierwszego -
    # dokleja sie do niego, dajac uszkodzony plik. Dlatego czyscimy bufor przed
    # finalnym zapisem. Przy zapisie na dysk (str) nie trzeba - otwarcie pliku
    # po nazwie samo kasuje poprzednia zawartosc.
    if not isinstance(sciezka, str):
        sciezka.seek(0)
        sciezka.truncate()
    wb.save(sciezka)
    return sciezka


def eksportuj_do_bajtow(df, raport):
    """Generuje raport Excel w PAMIECI i zwraca bajty (do pobrania w Streamlit).

    Tej samej logiki co eksportuj_do_excela uzywamy do zapisu na dysk lokalnie;
    tu kierujemy wynik do BytesIO (plik w pamieci) zamiast na dysk serwera,
    bo uzytkownik w przegladarce nie ma dostepu do dysku serwera."""
    bufor = io.BytesIO()                        # "plik" istniejacy tylko w pamieci
    eksportuj_do_excela(df, raport, sciezka=bufor)  # ta sama funkcja, inny cel zapisu
    bufor.seek(0)                               # przewin na poczatek przed odczytem
    return bufor.getvalue()                     # surowe bajty pliku xlsx


if __name__ == "__main__":
    # Test z terminala. Docelowo eksport wola Streamlit.
    from czyszczenie import wyczysc_dane   # import z DRUGIEGO pliku
    df, raport = wyczysc_dane()
    sciezka = eksportuj_do_excela(df, raport)
    print(f"Raport zapisany: {sciezka}")